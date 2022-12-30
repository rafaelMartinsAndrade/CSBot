import gui

#Imports necessários
import time
import threading
from queue import Queue
import json
import math as m
import urllib.parse
from collections import defaultdict
# from datetime import datetime

# Imports do xlsx
import pandas as pd
# import numpy as np
from openpyxl import workbook
from openpyxl import load_workbook


# import keyboard
# import random
# import win32api, win32con

#Imports do selenium
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoSuchWindowException
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.common.exceptions import StaleElementReferenceException

from pathlib import Path

OUTPUT_PATH = Path(__file__).parent
ASSETS_PATH = OUTPUT_PATH / Path("./assets")

def relative_to_assets(path: str) -> Path:
    return ASSETS_PATH / Path(path)

global ProfilePath
ProfilePath = Path(__file__).parent / Path("./SeleniumProfile")

def iniciarSessao():
	try:
		# global inicioScrapper
		# inicioScrapper = datetime.now()
		# inicioScrapper = inicioScrapper.strftime("%H:%M:%S")
		# print('Inicio: {0}, Fim: -'.format(inicioScrapper))

		#Variáveis de configuração da sessao
		driver_path = 'chromedriver'
		download_dir = "D:\\selenium"
		options = Options()
		options.add_argument("--headless")
		options.add_argument('--no-sandbox')
		options.add_experimental_option('excludeSwitches', ['enable-logging'])
		options.add_argument("--window-size=1000,800")
		options.add_argument('--allow-running-insecure-content')
		options.add_argument('--ignore-certificate-errors')
		# options.add_argument(r"user-data-dir={0}".format(ProfilePath))
		global sessao
		#Chama a função que abre a sessão com as configurações e o driver já prédefinidos
		sessao = webdriver.Chrome(options=options, executable_path=driver_path)
		#Redireciona a sessão para a url
		sessao.get(r"https://startupscanner.com/startups-data?per_page=1&page=1")
		logar()
	except:
		gui.atualizarLog('Não foi possível pesquisar as startup\'s')
		try:
			sessao.quit()
		except:
			pass
		return False

	# global fimScrapper
	# fimScrapper = datetime.now()
	# fimScrapper = fimScrapper.strftime("%H:%M:%S")
	# print('Inicio: {0}, Fim: {1}'.format(inicioScrapper, fimScrapper))

def logar():
	try:
		gui.atualizarLog('Verificando a quantidade de startup\'s!')
		element = WebDriverWait(sessao, 5).until(
			EC.element_to_be_clickable((By.XPATH, "//*[@id='emailInput']"))
		)
		inputEmail = sessao.find_element("xpath", "//*[@id='emailInput']")
		inputEmail.send_keys('rafaelmartinsdandrade@gmail.com')
		inputSenha = sessao.find_element("xpath", "//*[@id='passwordInput']")
		inputSenha.send_keys('senhaStartupScanner')
		btnLogin = sessao.find_element("xpath", "//*[@id='login-button']")
		btnLogin.click()
		processarJson()
	except TimeoutException:
		processarJson()
	except NoSuchElementException as e:
		gui.atualizarLog('Não foi possível pesquisar as startup\'s!')
		sessao.quit()
		return False

def processarJson():
	try:
		element = WebDriverWait(sessao, 5).until(
			EC.presence_of_element_located((By.XPATH, "/html/body/pre"))
		)
		preStartupsTemp = sessao.find_element("xpath", "/html/body/pre")
		jsonStartupsTemp = json.loads(preStartupsTemp.text)

		global arrStartups
		arrStartups = []

		global totalStartups
		totalStartups = jsonStartupsTemp['total']

		for i in range(1,(m.ceil(totalStartups/1000)+1)):
			gui.atualizarLog('Pesquisando {0} startup\'s! {1} de {2}'.format(totalStartups, i, m.ceil(totalStartups/1000)))
			sessao.get("https://startupscanner.com/startups-data?per_page={0}&page={1}".format(1000,i))
			element = WebDriverWait(sessao, 120).until(
				EC.presence_of_element_located((By.XPATH, "/html/body/pre"))
			)
			preStartups = sessao.find_element("xpath", "/html/body/pre")
			arrStartupsTemp = json.loads(preStartups.text)

			if "data" in arrStartupsTemp:
				for arrStartup in arrStartupsTemp["data"]:
					arrStartups.append(arrStartup["name"])
			else:
				break

		sessao.quit()
		gui.atualizarLog('As startup\'s foram carregadas!')
		verificarExcel()
	except TimeoutException:
		gui.atualizarLog('As startup\'s não carregaram!')
		return False
	except:
		gui.atualizarLog('Ocorreu um erro na pesquisa!')
		return False


def verificarExcel():
	try:
		excel = pd.read_excel(r"startups.xlsx",'Sheet1',header=1)
	except:
		gui.atualizarLog('Ocorreu um erro ao abrir o excel')
		return False
	try:
		excelStartups = excel["Nome da empresa"]
		excelLinkedin = excel["Linkedin"]
	except:
		gui.atualizarLog('O excel está desconfigurado!')
		return False

	try:
		global totalStartups
		global qStartupsPesquisa
		qStartupsPesquisa = Queue()

		gui.atualizarLog('Verificando 1 de {0}'.format(totalStartups))

		iStartupsExcel = 0
		for nomeStartup in arrStartups:
			gui.atualizarLog('Verificando {0} de {1}'.format(iStartupsExcel, totalStartups))
			if not excelStartups.str.contains(nomeStartup,regex=False).any():
				qStartupsPesquisa.put(nomeStartup)
			iStartupsExcel += 1
	except:
		gui.atualizarLog('Ocorreu um erro ao verificar as Startup\'s!')
		return False

	gui.atualizarLog('As startup\'s foram verificadas no excel!')
	time.sleep(1)
	
	if qStartupsPesquisa.qsize() == 0:
		gui.atualizarLog('Não foram encontradas novas startup\'s!')
		return True

	global dictStartupsExcel
	dictStartupsExcel = defaultdict(dict)

	iPesquisarStartups = 0
	nOldThreads = threading.active_count();
	global pesquisandoLinkedin
	pesquisandoLinkedin = True
	while not qStartupsPesquisa.empty():
		nThreads = threading.active_count()
		if nThreads < nOldThreads+5:
			threadPesquisarLinkedin = threading.Thread(target=pesquisarLinkedin, args=(), daemon=True)
			threadPesquisarLinkedin.start()
			iPesquisarStartups += 1
			gui.atualizarLog('Pesquisando os linkedin\'s! {0} de {1}'.format(iPesquisarStartups,iStartupsExcel))
		time.sleep(0.5)

	while pesquisandoLinkedin:
		time.sleep(0.5)

	if inserirExcel():
		gui.atualizarLog('{0} startup\'s foram inseridas no excel!'.format(len(dictStartupsExcel)))
		return True

	return False

def pesquisarLinkedin():
	try:
		strStartup = qStartupsPesquisa.get()

		#Variáveis de configuração da sessao
		driver_path = 'chromedriver'
		download_dir = "D:\\selenium"
		options = Options()
		options.add_argument("--headless")
		options.add_argument('--no-sandbox')
		options.add_experimental_option('excludeSwitches', ['enable-logging'])
		options.add_argument("--window-size=1000,800")
		options.add_argument('--allow-running-insecure-content')
		options.add_argument('--ignore-certificate-errors')
		# options.add_argument(r"user-data-dir={0}".format(ProfilePath))
		#Chama a função que abre a sessão com as configurações e o driver já prédefinidos
		sessao = webdriver.Chrome(options=options, executable_path=driver_path)
		#Redireciona a sessão para a url
		sessao.get(r"https://www.google.com/search?q={0}+empresa+linkedin".format(urllib.parse.quote_plus(strStartup)))
	except:
		gui.atualizarLog('Ocorreu um erro ao iniciar a pesquisa da startup {0}'.format(strStartup))
		qStartupsPesquisa.put(strStartup)
		return False
	try:
		element = WebDriverWait(sessao, 10).until(
			EC.presence_of_element_located((By.XPATH, "(//a[./h3[contains(text(), 'LinkedIn')] and (contains(@href, 'https://www.linkedin.com/company') or contains(@href, 'https://br.linkedin.com/company'))])[1]"))
		)
		try:
			dictStartupsExcel[strStartup] = ""
			linkLinkedin = sessao.find_element("xpath", "(//a[./h3[contains(text(), 'LinkedIn')] and (contains(@href, 'https://www.linkedin.com/company') or contains(@href, 'https://br.linkedin.com/company'))])[1]")
			urlLinkedin = linkLinkedin.get_attribute('href')
			dictStartupsExcel[strStartup] = urlLinkedin
		except:
			print('Ocorreu um erro ao processar o linkedin da startup {0}'.format(strStartup))
		if qStartupsPesquisa.empty():
			global pesquisandoLinkedin
			pesquisandoLinkedin = False
		sessao.quit()
		return True
	except:
		print('Ocorreu um erro ao pesquisar a startup {0}'.format(strStartup))
		# qStartupsPesquisa.put(strStartup)
		return False

def inserirExcel():
	global iStartupsInseridas
	global dictStartupsExcel

	iStartupsInseridas = 0

	try:
		i = 0
		excel = pd.ExcelFile("startups.xlsx")
		df = excel.parse('Sheet1', skiprows=2, index_col=1,header = None, na_values=['NA'])
	except:
		gui.atualizarLog('Ocorreu um erro ao abrir o excel')
		return False

	try:
		if not df.empty:
			arrExcelStartups = df[0]
			for startup in arrExcelStartups:
				try:
					m.isnan(float(startup))
					break
				except:
					i += 1
		wb = load_workbook("startups.xlsx")
		sheets = wb.sheetnames
		sheet1 = wb[sheets[0]]
		for startup in dictStartupsExcel:
			try:
				j = i+3
				sheet1.cell(row = j, column = 1).value = startup
				sheet1.cell(row = j, column = 4).value = dictStartupsExcel[startup]
				iStartupsInseridas += 1
				i+=1
			except:
				iStartupsInseridas -= 1
	except:
		gui.atualizarLog('Ocorreu um erro ao processar o excel')
		return False

	try:
		wb.save("startups.xlsx")
	except:
		gui.atualizarLog('Ocorreu um erro ao salvar o excel')
		return False

	return True

if __name__ == '__main__':
	iniciarSessao()